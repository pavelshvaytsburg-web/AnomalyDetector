import argparse
import os
import numpy as np
import pandas as pd

from modules.data_loader import load_data
from modules.model import (
    train_isolation_forest,
    detect_anomalies_scores,
    detect_anomalies_custom,
    build_results_dataframe
)
from modules.frequency_analysis import frequency_analysis
from modules.permutation_importance import permutation_importance_unsupervised
from modules.visualization import plot_anomaly_scores, plot_packets_vs_bytes

from openpyxl import Workbook
from openpyxl.styles import PatternFill


# =============================================
# РУССКИЕ НАЗВАНИЯ СТОЛБЦОВ
# =============================================
COLUMN_MAP = {
    "source_port": "Исходный порт",
    "destination_port": "Порт назначения",
    "protocol": "Протокол",
    "duration": "Длительность (сек)",
    "packet_count": "Кол-во пакетов",
    "bytes_sent": "Отправлено (байт)",
    "bytes_received": "Получено (байт)",
    "bytes_per_packet": "Средний размер пакета",
    "anomaly_score": "Оценка аномальности",
    "anomaly_label": "Метка модели",
    "freq_count": "Частота появления",
    "rarity_flag": "Редкость",
    "final_anomaly": "Итоговая аномалия",
    "comment": "Комментарий"
}


# =============================================
# ГЕНЕРАЦИЯ КОММЕНТАРИЯ
# =============================================
def generate_comment(row):
    reasons = []

    if row["anomaly_score"] < -0.05:
        reasons.append("низкая оценка anomaly_score")

    if row["rarity_flag"]:
        reasons.append("редкий тип трафика")

    if row["bytes_sent"] > 100000:
        reasons.append("большой объём исходящих данных")

    if row["bytes_per_packet"] > 300:
        reasons.append("крупный средний размер пакета")

    if row["duration"] > 10:
        reasons.append("длительное соединение")

    if not reasons:
        return "Параметры заметно отклоняются от нормы"

    return "; ".join(reasons)


# =============================================
# СОХРАНЕНИЕ EXCEL ОТЧЁТА
# =============================================
def save_excel_report(df: pd.DataFrame, output_path: str):

    df = df.copy()
    df["comment"] = df.apply(generate_comment, axis=1)

    df_ru = df.rename(columns=COLUMN_MAP)

    wb = Workbook()
    ws = wb.active
    ws.title = "Аномалии"

    # Заголовки
    ws.append(list(df_ru.columns))

    # Цвета
    red_fill = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

    # Строки
    for idx, row in df_ru.iterrows():
        ws.append(list(row.values))
        excel_row = idx + 2

        fill = red_fill if row["Итоговая аномалия"] == 1 else green_fill
        for cell in ws[excel_row]:
            cell.fill = fill

    wb.save(output_path)
    print(f"[INFO] Excel отчёт сохранён: {output_path}")


# =============================================
# ОСНОВНАЯ ФУНКЦИЯ
# =============================================
def main():
    parser = argparse.ArgumentParser(description="Unsupervised ML Anomaly Detector")
    parser.add_argument("--input", type=str, required=True)
    parser.add_argument("--output", type=str, default="output")
    args = parser.parse_args()

    os.makedirs(args.output, exist_ok=True)

    print(os.getcwd())
    print("Expecting file:", args.input)

    # 1. ЗАГРУЗКА
    df_features, X_scaled, scaler = load_data(args.input)
    print(f"[INFO] Загружено строк: {len(df_features)}")

    # 2. ОБУЧЕНИЕ IF
    model = train_isolation_forest(X_scaled)

    # 3. SCORE
    scores = detect_anomalies_scores(model, X_scaled)

    # 4. АНОМАЛИИ ПО THRESHOLD + IQR
    labels, final_thr, p_thr, iqr_thr = detect_anomalies_custom(scores)

    df_results = build_results_dataframe(df_features, scores, labels)

    # 5. ЧАСТОТНЫЙ АНАЛИЗ
    df_results = frequency_analysis(df_results)

    df_results["final_anomaly"] = (
        (df_results["anomaly_label"] == -1) &
        (df_results["rarity_flag"])
    ).astype(int)

    # 6. ВАЖНОСТЬ ПРИЗНАКОВ
    importance_df = permutation_importance_unsupervised(model, df_features)
    importance_path = os.path.join(args.output, "feature_importance.csv")
    importance_df.to_csv(importance_path, index=False)
    print(f"[INFO] Feature importance сохранён: {importance_path}")

    # 7. СОХРАНЕНИЕ CSV
    csv_path = os.path.join(args.output, "anomaly_results.csv")
    df_results.to_csv(csv_path, index=False)
    print(f"[INFO] CSV сохранён: {csv_path}")

    # 8. EXCEL ОТЧЁТ
    excel_path = os.path.join(args.output, "anomaly_report.xlsx")
    save_excel_report(df_results, excel_path)

    # 9. ВИЗУАЛИЗАЦИЯ
    hist_path = os.path.join(args.output, "anomaly_scores_hist.png")
    scatter_path = os.path.join(args.output, "packets_vs_bytes.png")

    plot_anomaly_scores(df_results, hist_path)
    plot_packets_vs_bytes(df_results, scatter_path)

    print(f"[INFO] Гистограмма сохранена: {hist_path}")
    print(f"[INFO] Scatter plot сохранён: {scatter_path}")

    print("[INFO] Работа завершена.")


if __name__ == "__main__":
    main()
